import streamlit as st
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO

# --- CONFIGURATION & STYLING ---
st.set_page_config(page_title="CSCEC Payment Dashboard", layout="wide", initial_sidebar_state="expanded")

# Colors
CSCEC_BLUE = "#00529B"
BACKGROUND_LIGHT = "#FFFFFF"
TEXT_DARK = "#1A1A1A"
BORDER_COLOR = "#E0E0E0"

# Files
CONFIG_FILE = "data/config.json"
LOGO_PATH = "assets/cscec.png"

# Custom CSS for Premium Look
st.markdown(f"""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;800&display=swap');
    
    html, body, [class*="css"] {{
        font-family: 'Inter', sans-serif;
    }}
    
    .stApp {{
        background-color: var(--background-color);
    }}
    
    /* Section Titles */
    .section-title {{
        color: {CSCEC_BLUE};
        font-size: 1.2rem;
        font-weight: 800;
        margin-bottom: 20px;
        padding-bottom: 8px;
        border-bottom: 2px solid {CSCEC_BLUE};
        text-transform: uppercase;
        letter-spacing: 1px;
    }}
    
    /* Labels */
    .cn-label {{
        color: var(--text-color);
        font-size: 0.8rem;
        font-weight: 600;
        margin-bottom: 2px;
    }}
    .en-label {{
        color: var(--text-color);
        font-size: 0.95rem;
        font-weight: 700;
        margin-bottom: 8px;
    }}
    
    /* Calculated Values */
    .calc-value {{
        background: var(--secondary-background-color);
        padding: 10px 15px;
        border-radius: 4px;
        border-left: 4px solid {CSCEC_BLUE};
        font-family: 'Courier New', monospace;
        font-size: 1.1rem;
        font-weight: 800;
        color: {CSCEC_BLUE};
        margin-bottom: 15px;
    }}
    
    /* Sidebar Styling */
    section[data-testid="stSidebar"] {{
        border-right: 1px solid var(--secondary-background-color);
    }}
    
    /* Hide default Streamlit elements */
    /* Premium Sidebar Navigation */
    .nav-item {{
        display: block;
        padding: 10px 15px;
        color: var(--text-color) !important;
        text-decoration: none !important;
        font-weight: 600;
        font-size: 0.9rem;
        border-radius: 6px;
        margin-bottom: 4px;
        transition: 0.2s ease all;
        border-left: 4px solid transparent;
    }}
    .nav-item:hover {{
        background-color: {CSCEC_BLUE};
        color: white !important;
        padding-left: 20px;
        border-left: 4px solid var(--background-color);
    }}
    
    #MainMenu {{ visibility: visible; }}
    footer {{ visibility: hidden; }}
    </style>
""", unsafe_allow_html=True)

# --- HELPER FUNCTIONS ---
def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, 'r') as f:
            return json.load(f)
    return {}

def save_config(config):
    os.makedirs(os.path.dirname(CONFIG_FILE), exist_ok=True)
    with open(CONFIG_FILE, 'w') as f:
        json.dump(config, f)

def export_to_excel(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Settlement Report"
    thin = Side(style='thin')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="00529B", end_color="00529B", fill_type="solid")
    calc_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 75
    ws.column_dimensions['C'].width = 25
    
    ws['B1'] = "CSCEC - Payment Voucher Settlement"
    ws['B1'].fill = header_fill
    ws['B1'].font = Font(bold=True, color="FFFFFF", size=12)
    
    row = 3
    labels = [
        ("1", "Initial accumulative supplier settlement amount", "1A"),
        ("", "Current period supplier settlement without VAT", "1B"),
        ("", "Current period other-payables under contract", "1C"),
        ("", "Current period other-deductions under contract", "1D"),
        ("", "Current period VAT of supplier settlement amount", "1E"),
        ("", "Current period settlement incl. VAT", "1G"),
        ("", "Ending accumulative supplier settlement amount", "1F"),
        ("", "", ""),
        ("2", "Advance payment under contract", "2A"),
        ("", "Initial deduction from advance payment", "2B"),
        ("", "Current period deduction from advance payment", "2C"),
        ("", "Ending deduction from advance payment", "2D"),
        ("", "Ending balance of advance payment", "2E"),
        ("", "", ""),
        ("3", "Ending accumulative amount payable", "3A"),
        ("", "", ""),
        ("4", "Initial balance of retention amount", "4A"),
        ("", "Current period retention money to be deducted", "4B"),
        ("", "Current period return of retention money", "4C"),
        ("", "Ending balance of retention money", "4D"),
        ("", "", ""),
        ("5", "Initial balance of temporary labour social insurance", "5A"),
        ("", "Current temporary labour to deduct", "5B"),
        ("", "Current return of temporary labour social insurance", "5C"),
        ("", "Ending balance of temporary labour social insurance", "5D"),
        ("", "", ""),
        ("6", "Initial accumulative WHT", "6A"),
        ("", "Current period WHT", "6B"),
        ("", "Ending accumulative WHT", "6C"),
        ("", "Settlement incl. VAT minus WHT", "6D"),
        ("", "", ""),
        ("8", "Initial other-deductions", "8A"),
        ("", "Current other-deductions", "8B"),
        ("", "Ending accumulative other-deductions", "8C"),
        ("", "", ""),
        ("12", "Initial accumulative contract social insurance", "12A"),
        ("", "Current period contract social insurance", "12B"),
        ("", "Ending accumulative contract social insurance", "12C"),
        ("", "", ""),
        ("7", "Initial accumulative reality amount paid", "7A"),
        ("", "Initial total accumulative paid amount", "7B"),
        ("", "", ""),
        ("9", "Net amount payable", "9A"),
        ("", "", ""),
        ("10", "Current period reality amount paid", "10A"),
        ("", "", ""),
        ("11", "Ending accumulative reality amount paid", "11A"),
        ("", "Ending total accumulative amount paid", "11B"),
    ]
    
    for section, desc, key in labels:
        if desc == "":
            row += 1
            continue
            
        ws[f'A{row}'] = section
        ws[f'B{row}'] = desc
        ws[f'C{row}'] = data.get(key, 0.0)
        ws[f'C{row}'].number_format = '#,##0.00'
        
        # highlighting logic
        if key in ["1F", "1G", "2D", "2E", "3A", "4D", "5D", "6C", "6D", "8C", "12C", "9A", "11A", "11B"]:
            ws[f'C{row}'].fill = calc_fill
            
        row += 1

    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)
    return virtual_workbook.getvalue()

def main():
    config = load_config()
    if 'config' not in st.session_state:
        st.session_state.config = config
    
    col_logo, col_title = st.columns([1, 15])
    with col_logo:
        try: st.image(LOGO_PATH, width=60)
        except: st.markdown(f"<b style='color:var(--text-color); font-size:14px; font-weight: 800;'>CSCEC</b>", unsafe_allow_html=True)
    with col_title:
        st.markdown(f"""
        <div style="background: var(--background-color); border: 2px solid var(--text-color); padding: 10px 20px;">
            <div style="color: var(--text-color); font-size: 18px; font-weight: 800; font-family: 'Microsoft YaHei', sans-serif;">Payment Voucher for Approval | 中国建筑管理表格</div>
            <div style="color: var(--text-color); opacity: 0.7; font-size: 12px; text-transform: uppercase; font-weight:600;">China State Construction Engineering Corporation</div>
        </div>
        """, unsafe_allow_html=True)

    def get_val(key, default=0.0): return st.session_state.config.get(key, default)
    def save_val(key, value): st.session_state.config[key] = value

    def num_input(zh, en, key, default_val=0.0):
        field_id = f"f_{key}"
        st.markdown(f'<div id="{field_id}" class="cn-label">{zh}</div><div class="en-label">{en}</div>', unsafe_allow_html=True)
        val = st.text_input("hidden", value=str(get_val(key, default_val)), key=f"n_{key}", label_visibility="collapsed")
        try: val = float(val or 0)
        except: val = default_val
        save_val(key, val)
        st.markdown('<div style="margin-bottom:8px;"></div>', unsafe_allow_html=True)
        return val

    def calc_display(zh, en, value):
        st.markdown(f'<div class="cn-label">{zh}</div><div class="en-label">{en}</div>', unsafe_allow_html=True)
        st.markdown(f'<div class="calc-value">EGP {value:,.2f}</div>', unsafe_allow_html=True)

    # --- CALCULATION PHASE ---
    def get_raw_val(key, default=0.0): 
        try: return float(st.session_state.config.get(key, default) or 0)
        except: return default
    
    c_1A = get_raw_val("val_1A", 1715291.2)
    c_1B = get_raw_val("val_1B", 0.0)
    c_1C = get_raw_val("val_1C", 0.0)
    c_1D = get_raw_val("val_1D", 0.0)
    c_vat_rate = get_raw_val("vat_rate", 14)
    c_1E = round(c_1B * c_vat_rate / 100, 2)
    c_1G = c_1B + c_1E
    c_1F = c_1A + c_1B + c_1C - c_1D + c_1E
    
    c_2A = get_raw_val("val_2A", 277500.0)
    c_2B = get_raw_val("val_2B", 256358.13)
    c_2C = get_raw_val("val_2C", 935.25)
    c_2D = c_2B + c_2C
    c_2E = c_2A - c_2D
    c_3A = c_1F - c_2D
    
    c_ret_rate = get_raw_val("ret_rate", 0)
    c_4A = get_raw_val("val_4A", 0.0)
    c_4B = round(c_1B * c_ret_rate / 100, 2)
    c_4C = get_raw_val("val_4C", 0.0)
    c_4D = c_4A + c_4B - c_4C
    
    c_temp_rate = get_raw_val("temp_rate", 0)
    c_5A = get_raw_val("val_5A", 0.0)
    c_5B = round(c_1B * c_temp_rate / 100, 2)
    c_5C = get_raw_val("val_5C", 0.0)
    c_5D = c_5A + c_5B - c_5C
    
    c_wht_rate = get_raw_val("wht_rate", 0)
    c_6A = get_raw_val("val_6A", 0.0)
    c_6B = round(c_1B * c_wht_rate / 100, 2)
    c_6C = c_6A + c_6B
    c_6D = c_1G - c_6B
    
    c_oth_rate = get_raw_val("oth_rate", 0)
    c_8A = get_raw_val("val_8A", 0.0)
    c_8B = round(c_1B * c_oth_rate / 100, 2)
    c_8C = c_8A + c_8B
    
    c_soc_rate = get_raw_val("soc_rate", 0)
    c_12A = get_raw_val("val_12A", 0.0)
    c_12B = round(c_1B * c_soc_rate / 100, 2)
    c_12C = c_12A + c_12B
    
    c_total_deductions = c_4B + c_5B + c_6B + c_8B + c_12B
    c_7A = get_raw_val("val_7A", 1302933.73)
    c_7B = c_7A + c_6A
    c_9A = c_3A - c_4D - c_5D - c_6C - c_7A - c_8C - c_12C
    c_10A = get_raw_val("val_10A", 0.0)
    c_11A = c_7A + c_10A + c_2A
    c_11B = c_11A + c_6C

    # --- SIDEBAR: NAVIGATOR & SUMMARY ---
    st.sidebar.markdown(f"""
    <div style="padding: 10px; background: {CSCEC_BLUE}; color: white; font-weight: 800; text-align: center; margin-bottom: 20px;">
        AUDIT DASHBOARD | 审计板
    </div>
    """, unsafe_allow_html=True)

    st.sidebar.markdown("### 📊 KEY TOTALS")
    with st.sidebar.container(border=True):
        st.markdown(f"**Total Deductions:**<br/><span style='color:{CSCEC_BLUE}; font-size:1.1em; font-weight:800;'>EGP {c_total_deductions:,.2f}</span>", unsafe_allow_html=True)
        st.markdown(f"**Net Payable (9A):**<br/><span style='color:{CSCEC_BLUE}; font-size:1.1em; font-weight:800;'>EGP {c_9A:,.2f}</span>", unsafe_allow_html=True)
        st.markdown(f"**Reality Payment (11A):**<br/><span style='color:{CSCEC_BLUE}; font-size:1.1em; font-weight:800;'>EGP {c_11A:,.2f}</span>", unsafe_allow_html=True)
    
    st.sidebar.markdown(f"""
        <div style="margin-top:20px;"></div>
        <div style="color:#888; font-size:0.75rem; font-weight:700; text-transform:uppercase; margin-bottom:10px; padding-left:15px;">Main Overview</div>
        <a class="nav-item" href="#f_1B">Settlement Amount</a>
        <a class="nav-item" href="#f_1E">VAT Details</a>
        <a class="nav-item" href="#f_2A">Advance Payment</a>
        <a class="nav-item" href="#f_3A">Accumulative Payable</a>
        <div style="margin-top:20px; color:#888; font-size:0.75rem; font-weight:700; text-transform:uppercase; margin-bottom:10px; padding-left:15px;">Deductions</div>
        <a class="nav-item" href="#f_4B">Retention Money</a>
        <a class="nav-item" href="#f_6B">Withholding Tax (WHT)</a>
        <a class="nav-item" href="#f_5B">Temp Labour Social</a>
        <a class="nav-item" href="#f_12B">Contract Social Insurance</a>
        <a class="nav-item" href="#f_8B">Other Deductions</a>
        <div style="margin-top:20px; color:#888; font-size:0.75rem; font-weight:700; text-transform:uppercase; margin-bottom:10px; padding-left:15px;">Final Audit</div>
        <a class="nav-item" href="#f_9A">Final Net Result</a>
    """, unsafe_allow_html=True)

    # 1. SUPPLIER SETTLEMENT
    st.markdown('<div id="f_1B"></div>', unsafe_allow_html=True)
    with st.container(border=True):
        st.markdown('<div class="section-title">Step 1: 供应商结算 | Supplier Settlement</div>', unsafe_allow_html=True)
        col1, col2 = st.columns(2)
        with col1:
            val_1A = num_input("期初累计供应商结算金额", "Initial accumulative supplier settlement amount", "val_1A", 1715291.2)
            val_1B = num_input("本期供应商结算（不含税）", "Current period supplier settlement (w/o VAT)", "val_1B", 0.0)
            val_1C = num_input("本期合同项下其他应付", "Current period other-payables under contract", "val_1C", 0.0)
        with col2:
            val_1D = num_input("本期合同项下其他扣款", "Current period other-deductions under contract", "val_1D", 0.0)
            st.markdown(f'<div class="cn-label">增值税率</div><div class="en-label">VAT Rate</div>', unsafe_allow_html=True)
            vat_options = [0, 14, 9, 5, 10]
            vat_val = get_val("vat_rate", 14)
            vat_idx = vat_options.index(vat_val) if vat_val in vat_options else 1
            vat_rate = st.selectbox("VAT Rate", vat_options, index=vat_idx, key="s_vat", label_visibility="collapsed")
            save_val("vat_rate", vat_rate)
            st.markdown('<div style="margin-bottom:8px;"></div>', unsafe_allow_html=True)
            st.markdown('<div id="f_1E"></div>', unsafe_allow_html=True)
            calc_display("本期增值税额", "Current period VAT amount", c_1E)
            calc_display("本期供应商结算(含税)", "Current period settlement incl. VAT", c_1G)
        st.divider()
        calc_display("期末累计供应商结算金额", "Ending accumulative supplier settlement amount", c_1F)

    # 2. ADVANCE PAYMENT
    st.markdown('<div id="f_2A"></div>', unsafe_allow_html=True)
    with st.container(border=True):
        st.markdown('<div class="section-title">Step 2: 预付款 | Advance Payment</div>', unsafe_allow_html=True)
        col1, col2 = st.columns(2)
        with col1:
            val_2A = num_input("合同项下预付款", "Advance payment under contract", "val_2A", 277500.0)
            val_2B = num_input("期初预付款扣除金额", "Initial deduction from advance payment", "val_2B", 256358.13)
        with col2:
            val_2C = num_input("本期预付款扣除金额", "Current period deduction from advance payment", "val_2C", 935.25)
            calc_display("期末预付款扣除金额", "Ending deduction from advance payment", c_2D)
            calc_display("期末预付款余额", "Ending balance of advance payment", c_2E)

    # 3. ENDING ACCUMULATIVE AMOUNT PAYABLE
    st.markdown('<div id="f_3A"></div>', unsafe_allow_html=True)
    with st.container(border=True):
        st.markdown('<div class="section-title">Step 3: 累计应付金额 | Ending Accumulative Payable</div>', unsafe_allow_html=True)
        calc_display("期末累计应付金额", "Ending accumulative amount payable", c_3A)

    # DEDUCTIONS
    with st.container(border=True):
        st.markdown('<div class="section-title">Step 4-8: 扣款项目 | Deductions</div>', unsafe_allow_html=True)
        
        st.markdown('<div id="f_4B"></div>', unsafe_allow_html=True)
        with st.expander("🔗 4. 保留金 | RETENTION MONEY", expanded=False):
            col1, col2 = st.columns(2)
            with col1:
                val_4A = num_input("期初保留金余额", "Initial balance of retention amount", "val_4A", 0.0)
                st.markdown(f'<div class="cn-label">扣除比例</div><div class="en-label">Rate</div>', unsafe_allow_html=True)
                ret_options = [0, 3, 5, 10]
                ret_val = get_val("ret_rate", 0)
                ret_idx = ret_options.index(ret_val) if ret_val in ret_options else 0
                retention_rate = st.selectbox("Retention Rate", ret_options, index=ret_idx, key="s_ret", label_visibility="collapsed")
                save_val("ret_rate", retention_rate)
            with col2:
                calc_display("本期应扣保留金", "Current period retention money to be deducted", c_4B)
                val_4C = num_input("本期退回保留金", "Current period return of retention money", "val_4C", 0.0)
            st.divider()
            calc_display("期末保留金余额", "Ending balance of retention money", c_4D)

        st.markdown('<div id="f_5B"></div>', unsafe_allow_html=True)
        with st.expander("🔗 5. 临时劳务费社保 | TEMP LABOUR SOCIAL", expanded=False):
            col1, col2 = st.columns(2)
            with col1:
                val_5A = num_input("期初临时劳务费社保余额", "Initial balance of temporary labour social insurance", "val_5A", 0.0)
                st.markdown(f'<div class="cn-label">扣除比例</div><div class="en-label">Rate</div>', unsafe_allow_html=True)
                temp_options = [0, 0.45]
                temp_val = get_val("temp_rate", 0)
                temp_idx = temp_options.index(temp_val) if temp_val in temp_options else 0
                temp_rate = st.selectbox("Temp Rate", temp_options, index=temp_idx, key="s_temp", label_visibility="collapsed")
                save_val("temp_rate", temp_rate)
            with col2:
                calc_display("本期应扣临时劳务费", "Current temporary labour to deduct", c_5B)
                val_5C = num_input("本期退回临时劳务费", "Current return of temporary labour social insurance", "val_5C", 0.0)
            st.divider()
            calc_display("期末临时劳务费社保余额", "Ending balance of temporary labour social insurance", c_5D)

        st.markdown('<div id="f_6B"></div>', unsafe_allow_html=True)
        with st.expander("🔗 6. 预提税 | WITHHOLDING TAX", expanded=False):
            col1, col2 = st.columns(2)
            with col1:
                val_6A = num_input("期初累计预提税额", "Initial accumulative WHT", "val_6A", 0.0)
                st.markdown(f'<div class="cn-label">预提税率</div><div class="en-label">Rate</div>', unsafe_allow_html=True)
                wht_options = [0, 1, 3, 5]
                wht_val = get_val("wht_rate", 0)
                wht_idx = wht_options.index(wht_val) if wht_val in wht_options else 0
                wht_rate = st.selectbox("WHT Rate", wht_options, index=wht_idx, key="s_wht", label_visibility="collapsed")
                save_val("wht_rate", wht_rate)
            with col2:
                calc_display("本期预提税", "Current period WHT", c_6B)
                calc_display("期末累计预提税额", "Ending accumulative WHT", c_6C)
            st.divider()
            calc_display("决算金额含税减去预提税", "Settlement incl. VAT minus WHT", c_6D)

        st.markdown('<div id="f_8B"></div>', unsafe_allow_html=True)
        st.markdown('<div id="f_12B"></div>', unsafe_allow_html=True)
        with st.expander("🔗 8 & 12. 其他扣款与合同社保 | OTHERS & CONTRACT SOCIAL", expanded=False):
            col1, col2 = st.columns(2)
            with col1:
                val_8A = num_input("期初其他扣款", "Initial other-deductions", "val_8A", 0.0)
                st.markdown(f'<div class="cn-label">扣除费率</div><div class="en-label">Other bundle Rate</div>', unsafe_allow_html=True)
                oth_options = [0, 1, 2, 3, 5]
                oth_val = get_val("oth_rate", 0)
                oth_idx = oth_options.index(oth_val) if oth_val in oth_options else 0
                oth_rate = st.selectbox("Other Bundle Rate", oth_options, index=oth_idx, key="s_oth", label_visibility="collapsed")
                save_val("oth_rate", oth_rate)
            with col2:
                calc_display("本期其他扣款", "Current other-deductions", c_8B)
                val_8C = num_input("期末累计其他扣款", "Ending accumulative other-deductions", "val_8C", 0.0)
            st.divider()
            col3, col4 = st.columns(2)
            with col3:
                val_12A = num_input("期初累计合同项下社保", "Initial accumulative contract social insurance", "val_12A", 0.0)
                st.markdown(f'<div class="cn-label">社保费率</div><div class="en-label">Social Insurance Rate</div>', unsafe_allow_html=True)
                soc_options = [0, 3.6, 11.86, 3.3]
                soc_val = get_val("soc_rate", 0)
                soc_idx = soc_options.index(soc_val) if soc_val in soc_options else 0
                soc_rate = st.selectbox("Social Rate Box", soc_options, index=soc_idx, key="s_soc", label_visibility="collapsed")
                save_val("soc_rate", soc_rate)
            with col4:
                calc_display("本期合同项下社保", "Current period contract social insurance", c_12B)
                calc_display("期末累计合同项下社保", "Ending accumulative contract social insurance", c_12C)
        st.markdown(f"""
        <div style="background:var(--secondary-background-color); padding:10px; border:2px solid {CSCEC_BLUE}; margin-top:20px; display:flex; justify-content:space-between; align-items:center;">
            <div class="en-label" style="margin-bottom:0; color:var(--text-color);">Total Deductions | 扣款合计</div>
            <div style="font-family:monospace; font-size:16px; font-weight:800; color:{CSCEC_BLUE};">EGP {c_total_deductions:,.2f}</div>
        </div>
        """, unsafe_allow_html=True)

    # 7. ACCUMULATIVE PAID
    with st.container(border=True):
        st.markdown('<div class="section-title">Step 7: 累计已付 | Accumulative Paid</div>', unsafe_allow_html=True)
        col1, col2 = st.columns(2)
        with col1:
            val_7A = num_input("期初累计实际支付金额", "Initial accumulative reality amount paid", "val_7A", 1302933.73)
        with col2:
            calc_display("期初累计总支付金额", "Initial total accumulative paid amount", c_7B)

    # 9, 10 & 11. ENDING PAID
    st.markdown('<div id="f_9A"></div>', unsafe_allow_html=True)
    with st.container(border=True):
        st.markdown('<div class="section-title">Step 9-11: 实际与期末合计 | Realities & Finals</div>', unsafe_allow_html=True)
        col1, col2, col3 = st.columns(3)
        with col1:
            calc_display("应付款项净额", "Net amount payable", c_9A)
        with col2:
            st.markdown('<div style="height: 1px;"></div>', unsafe_allow_html=True)
            val_10A = num_input("本期实际付款", "Current period Reality amount paid", "val_10A", 0.0)
        with col3:
            st.markdown('<div style="height: 1px;"></div>', unsafe_allow_html=True)
            calc_display("期末累计实际已付款", "Ending accumulative reality amount paid", c_11A)
        st.divider()
        calc_display("期末累计已付款项合计", "Ending total accumulative amount paid", c_11B)

    st.markdown('<div style="margin-top: 25px;"></div>', unsafe_allow_html=True)
    col_space, col_save, col_exp = st.columns([3, 1, 1])
    with col_save:
        if st.button("Save Current", type="secondary", use_container_width=True):
            save_config(st.session_state.config)
            st.success("Saved!")
    with col_exp:
        data = {
            "1A": val_1A, "1B": val_1B, "1C": val_1C, "1D": val_1D, "1E": c_1E, "1F": c_1F, "1G": c_1G,
            "2A": val_2A, "2B": val_2B, "2C": val_2C, "2D": c_2D, "2E": c_2E,
            "3A": c_3A, "4A": val_4A, "4B": c_4B, "4C": val_4C, "4D": c_4D,
            "5A": val_5A, "5B": c_5B, "5C": val_5C, "5D": c_5D,
            "6A": val_6A, "6B": c_6B, "6C": c_6C, "6D": c_6D,
            "8A": val_8A, "8B": c_8B, "8C": c_8C,
            "12A": val_12A, "12B": c_12B, "12C": c_12C,
            "9A": c_9A, "10A": val_10A, "11A": c_11A, "11B": c_11B, "7A": val_7A, "7B": c_7B
        }
        excel_data = export_to_excel(data)
        st.download_button(
            label="Download Excel",
            data=excel_data,
            file_name=f"CSCEC_Settlement_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True
        )
    save_config(st.session_state.config)

if __name__ == "__main__":
    main()
